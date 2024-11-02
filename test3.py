import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import re

# 設定 Excel 檔案路徑
file_path = "C:\\Users\\User\\OneDrive - 逢甲大學\\文件\\GitHub\\group_7\\database2.xlsx"  # 課程表文件

# 開啟 Excel 檔案
workbook = openpyxl.load_workbook(file_path)
course_sheet = workbook["課程"]
student_sheet = workbook["學生"]

# GUI 設定
window = tk.Tk()
window.title("課程表顯示")
window.geometry("1200x600")

# 定義輸入框檢查學號格式
def validate_id_format(student_id):
    return bool(re.match(r"^D\d{7}$", student_id))

# 加課程到學生課表
def add_course_to_student(course_name, schedule, student_id):
    # 檢查學號格式
    if not validate_id_format(student_id):
        messagebox.showerror("錯誤", "學號格式錯誤。請輸入Dxxxxxxx格式的學號。")
        return

    # 查找學生個人課表路徑
    for row in student_sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == student_id:
            student_schedule_path = row[2]  # 個人課表路徑
            break
    else:
        messagebox.showerror("錯誤", "找不到該學號的學生。")
        return

    # 解析時間和星期
    day, time = parse_schedule(schedule)
    
    if not day or not time:
        messagebox.showerror("錯誤", "無效的時間或星期格式。")
        return

    # 開啟學生個人課表文件
    student_schedule_workbook = openpyxl.load_workbook(student_schedule_path)
    student_schedule_sheet = student_schedule_workbook.active

    # 將課程加到對應的時間與星期
    row = time_to_row(time)
    col = day_to_column(day)
    if row > 0 and col > 0:
        student_schedule_sheet.cell(row=row, column=col, value=course_name)

        # 保存檔案
        student_schedule_workbook.save(student_schedule_path)
        messagebox.showinfo("成功", f"課程 {course_name} 已添加到 {student_id} 的課表中。")
    else:
        messagebox.showerror("錯誤", "無效的時間或星期。")

# 輔助函數：解析「星期」和「時間」
def parse_schedule(schedule):
    match = re.match(r"^(星期[一二三四五])\s+(\d{2}:\d{2}-\d{2}:\d{2})$", schedule)
    if match:
        return match.group(1), match.group(2)
    return None, None

# 輔助函數：時間和星期對應的行列 (假設表格已知行列)
def time_to_row(time_str):
    time_map = {
        "08:00-10:00": 2,
        "10:00-12:00": 3,
        "13:00-15:00": 4,
        # 添加其他時間段...
    }
    return time_map.get(time_str, 0)

def day_to_column(day_str):
    day_map = {
        "星期一": 2,
        "星期二": 3,
        "星期三": 4,
        "星期四": 5,
        "星期五": 6,
    }
    return day_map.get(day_str, 0)

# 創建 Treeview 控件來顯示資料
tree = ttk.Treeview(window, show="headings")
tree.grid(row=0, column=1, sticky="nsew")

# 設置滾動條
x_scroll = tk.Scrollbar(window, orient="horizontal", command=tree.xview)
x_scroll.grid(row=1, column=1, sticky="ew")
y_scroll = tk.Scrollbar(window, orient="vertical", command=tree.yview)
y_scroll.grid(row=0, column=2, sticky="ns")

tree.configure(xscrollcommand=x_scroll.set, yscrollcommand=y_scroll.set)

# 讀取標題列 (第一列) 並設置 Treeview 的欄位
headers = [cell.value for cell in course_sheet[1]]
tree["columns"] = [f"col{i}" for i in range(len(headers))]  # 為每一列創建唯一的 ID

for i, header in enumerate(headers):
    tree.heading(f"col{i}", text=header)  # 使用唯一 ID 設置表頭
    tree.column(f"col{i}", anchor="center", minwidth=100)  # 設置列的對齊方式與最小寬度

# 讀取並顯示每一行內容，只顯示到最後有資料的行
for i, row in enumerate(course_sheet.iter_rows(min_row=2, values_only=True), start=2):
    if any(row):  # 確保當前行有資料才顯示
        course_name = row[0]
        schedule = row[2]  # 課程時間和星期的資訊

        # 在課程名稱前面放置輸入框
        student_id_entry = ttk.Entry(window)
        student_id_entry.grid(row=i, column=0)  # 輸入框放在第0列

        # 顯示課程資料
        tree.insert("", "end", values=row)

        # 加入課程按鈕
        add_button = ttk.Button(window, text="加入課程",
                                command=lambda c=course_name, s=schedule, e=student_id_entry: add_course_to_student(c, s, e.get()))
        add_button.grid(row=i, column=1)  # 按鈕放在第1列

# 自動調整列寬
window.grid_rowconfigure(0, weight=1)
window.grid_columnconfigure(0, weight=1)

# 啟動 GUI
window.mainloop()
