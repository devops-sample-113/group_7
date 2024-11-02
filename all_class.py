import tkinter as tk
from tkinter import ttk
import openpyxl

# 設定 GUI
window = tk.Tk()
window.title("課程表顯示")
window.geometry("1200x600")

# Excel 檔案路徑
file_path = "C:\\Users\\User\\OneDrive - 逢甲大學\\文件\\GitHub\\group_7\\database2.xlsx"

# 開啟 Excel 檔案
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook["課程"]

# 找到最後一行（包含資料的行數）
last_row = worksheet.max_row

# 創建 Treeview 控件來顯示資料
tree = ttk.Treeview(window, show="headings")
tree.grid(row=0, column=0, sticky="nsew")

# 設置滾動條
x_scroll = tk.Scrollbar(window, orient="horizontal", command=tree.xview)
x_scroll.grid(row=1, column=0, sticky="ew")
y_scroll = tk.Scrollbar(window, orient="vertical", command=tree.yview)
y_scroll.grid(row=0, column=1, sticky="ns")

tree.configure(xscrollcommand=x_scroll.set, yscrollcommand=y_scroll.set)

# 讀取標題列 (第一列) 並設置 Treeview 的欄位
headers = [cell.value for cell in worksheet[1]]
tree["columns"] = [f"col{i}" for i in range(len(headers))]  # 為每一列創建唯一的 ID

for i, header in enumerate(headers):
    tree.heading(f"col{i}", text=header)  # 使用唯一 ID 設置表頭
    tree.column(f"col{i}", anchor="center", minwidth=100)  # 設置列的對齊方式與最小寬度

# 讀取並顯示每一行內容，只顯示到最後有資料的行
for row in worksheet.iter_rows(min_row=2, max_row=last_row, values_only=True):
    if any(row):  # 確保當前行有資料才顯示
        tree.insert("", "end", values=row)

# 自動調整列寬
window.grid_rowconfigure(0, weight=1)
window.grid_columnconfigure(0, weight=1)

window.mainloop()
