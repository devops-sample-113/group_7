from tkinter import *
import tkinter as tk
import openpyxl
import subprocess
import os
import sys
from tkinter import messagebox
import tkinter.font as tkFont

def restart_app():
    Swindow.destroy()  # 關閉當前主視窗
    python = sys.executable  # 取得 Python 執行檔路徑
    subprocess.Popen([python] + sys.argv)  
    subprocess

##顯示課程資訊頁面
def show_course_details(value):
    #取得輸入資料並前往查詢
    with open("course.txt","w") as file:
        file.write(str(value))

    subprocess.Popen(["python", "course_detail.py"])  # 執行第二個程式

Swindow = Tk()
Swindow.title("選課系統")
Swindow.geometry("1500x750")

# 設置 Canvas 和 Scrollbar
canvas = Canvas(Swindow, highlightthickness=0)
scrollbar = Scrollbar(Swindow, orient="vertical", command=canvas.yview)
scrollbar.pack(side="right", fill="y")
canvas.configure(yscrollcommand=scrollbar.set)
canvas.place(x=10, y=130, width=1480, height=650)

# 新增框架用於承載內容
content_frame = Frame(canvas, width=1480, height=650)
canvas.create_window((0, 0), window=content_frame, anchor="nw")

# 用於顯示課程資料的子框架    ##查
#data_frame = Frame(content_frame, width=1480, height=650)
#data_frame.grid(row=1, column=0, columnspan=6, sticky="nsew")

# 當視窗大小變動時，動態調整 canvas 高度
def on_configure(event):
    canvas.configure(scrollregion=canvas.bbox("all"))
    # 根據視窗高度調整 canvas 的高度
    new_height = Swindow.winfo_height() - 130  # 130 是 canvas 的 y 位置
    canvas.place_configure(height=new_height)

# 綁定視窗大小變動事件
Swindow.bind("<Configure>", on_configure)

def on_configure(event):
    canvas.configure(scrollregion=canvas.bbox("all"))

content_frame.bind("<Configure>", on_configure)

def on_mouse_wheel(event):
    canvas.yview_scroll(-1 if event.delta > 0 else 1, "units")

canvas.bind_all("<MouseWheel>", on_mouse_wheel)  # Windows 和 Mac OS
#canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))  # Linux 向上
#canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))  # Linux 向下

# 開啟課表頁面
def open_new_window():
    subprocess.Popen(["python", "schedule_search.py"])  # 執行第二個程式

button_pop = Button(Swindow, text="課表頁面", command=open_new_window)
button_pop.place(x=100, y=10)

button_restart = Button(Swindow, text="重新載入首頁", command=restart_app)
button_restart.place(x=10, y=10)

headers = ["課程名稱", "課程代碼", "開課時間", "上課地點", "授課教授", "加退選框", "確認目前餘額"]
# 在 content_frame 中加入標題行
#for col_idx, header in enumerate(headers):
#    label = tk.Label(content_frame, text=header, borderwidth=1, relief="solid", padx=5, pady=5, bg="lightgray")
#    label.grid(row=0, column=col_idx, sticky="nsew", padx=2, pady=2)

# 開啟 Excel 文件
path = '資料庫.xlsx'
workbook = openpyxl.load_workbook(path)
worksheet_courses = workbook["課程"]
worksheet_students = workbook["學生"]

# 初始化 all_course 字典
all_course = {}
for row in worksheet_courses.iter_rows(min_row=2, values_only=True):
    course_name, course_code, course_time, location, professor, instructor_ID, ta_1, taID_1, ta_2, taID_2, syllabus, grading, enrollment_Limit, remaining_spots, credit = row[:15]
    all_course[course_code] = {
        "課程名稱": course_name,
        "開課時間": course_time,
        "上課地點": location,
        "授課教授": professor,
        "授課教師證號": instructor_ID,
        "課堂助教1": ta_1,
        "助教證號1": taID_1,
        "課堂助教2": ta_2,
        "助教證號2": taID_2,
        "課程大綱": syllabus,
        "評分方式": grading,
        "修課人數上限": enrollment_Limit,
        "目前可修課人數餘額": remaining_spots,
        "學分": credit
    }

# 查詢學號
def search(id):
    for row in worksheet_students.iter_rows(min_row=2, values_only=True):
        name, id_, schedule_path = row[:3]
        if id_ == id:
            return f"個人課表/{id}.xlsx"
    return None

# 取得課程學分
def get_course_credit(course_code):
    return all_course.get(course_code, {}).get("學分")

# 計算總學分
def calculate_total_credits(schedule_path):
    total_credits = 0
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active
    added_courses = set()

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if cell and cell not in added_courses:
                course_credit = get_course_credit(cell)
                if course_credit:
                    total_credits += course_credit
                    added_courses.add(cell)
    return total_credits

#檢查是否已含有相同名稱課程
def samecourse_existing(schedule_path, course_code):
    course_name = all_course[course_code]["課程名稱"]
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
        for cell in row:
            if cell and cell != course_code:  # 確保排除當前課程代碼
                # 取得該課程代碼的名稱
                existing_course_name = all_course.get(cell, {}).get('課程名稱')
                if existing_course_name == course_name:
                    return True
    return False

# 檢查課程是否在學生課表中
def is_course_in_schedule(schedule_path, course_code):
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if cell == course_code:
                return True
    return False

# 將課程時間對應至課表
def map_course_time_to_schedule(day, time_range):
    day_mapping = {"星期一": 2, "星期二": 3, "星期三": 4, "星期四": 5, "星期五": 6}
    start_time, _ = time_range.split('-')
    col_num = day_mapping[day]

    time_mapping = {
        "08:00": (2, 3), "09:00": (3, 4), "10:00": (4, 5),
        "11:00": (5, 6), "12:00": (6, 7), "13:00": (7, 8),
        "14:00": (8, 9), "15:00": (9, 10)
    }
    return col_num, *time_mapping.get(start_time, (None, None))

# 檢查課程是否衝堂
def check_schedule_conflict(schedule_path, col_num, start_row, end_row):
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active

    # 檢查開始與結束行的位置是否已經有課程
    if worksheet.cell(row=start_row, column=col_num).value or worksheet.cell(row=end_row, column=col_num).value:
        return True
    return False

# 獲取課程剩餘名額
def get_course_remaining_spots(course_code):
    return all_course.get(course_code, {}).get("目前可修課人數餘額")

# 新增課程至課表
def add_course_to_schedule(schedule_path, course_code, col_num, start_row, end_row):
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active
    
    worksheet.cell(row=start_row, column=col_num).value = course_code
    worksheet.cell(row=end_row, column=col_num).value = course_code
    workbook.save(schedule_path)

# 更新課程剩餘名額
def update_course_remaining_spots(course_code, new_remaining_spots):
    if course_code in all_course:
        all_course[course_code]["目前可修課人數餘額"] = new_remaining_spots
        for row in worksheet_courses.iter_rows(min_row=2, values_only=False):
            if row[1].value == course_code:
                row[13].value = new_remaining_spots
                workbook.save(path)
                break

# 顯示課表
def display_schedule(schedule_path):
    schedule_window = Toplevel(Swindow)
    schedule_window.title("個人課表")
    schedule_window.geometry("600x400")

    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True)):
        for col_idx, value in enumerate(row):
            label = tk.Label(schedule_window, text=value if value else "", borderwidth=1, relief="solid", padx=5, pady=5)
            label.grid(row=row_idx, column=col_idx, sticky="nsew", padx=2, pady=2)

    Button(schedule_window, text="關閉", command=schedule_window.destroy).grid(row=row_idx + 1, column=0, columnspan=len(row), pady=10)

# 退選課程
def drop_course_from_schedule(schedule_path, course_code):
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active
    found = False
    for row in worksheet.iter_rows(min_row=2, values_only=False):
        for cell in row:
            if cell.value == course_code:
                cell.value = None  # 刪除該課程
                found = True
                break
    if not found:
        messagebox.showinfo("退選失敗", f"課表中並沒有{all_course[course_code]['課程名稱']}{course_code}，退選失敗")
        return False
    total_credits = calculate_total_credits(schedule_path)
    course_credit = get_course_credit(course_code)
    if total_credits - course_credit < 9:
        messagebox.showerror("低於學分下限", "退選失敗，學分低於 9 學分！")
        return False   
    remaining_spots = get_course_remaining_spots(course_code)
    if remaining_spots is not None:
        update_course_remaining_spots(course_code, remaining_spots + 1)       
    workbook.save(schedule_path)
    messagebox.showinfo("退選成功", f"{all_course[course_code]['課程名稱']}{course_code}，退選成功")
    return True

###########################################搜尋

course_keyword = ""
number_keyword = ""
room_keyword = ""
professor_keyword = ""
week_keyword = ""
time_keyword = 0
week_options = ["", "一", "二", "三", "四", "五"]
time_options = ["", "08:00-09:00", "09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00", "16:00-17:00"]
time_number = [0, 8, 9, 10, 11, 12, 13, 14, 15, 16]

def update_keywords():
    global professor_keyword, course_keyword, number_keyword, room_keyword, week_keyword, time_keyword  # 讓這兩個變數可以在函數外部使用並更新
    course_keyword = course_entry.get().strip()
    number_keyword = number_entry.get().strip()
    room_keyword = room_entry.get().strip()
    professor_keyword = professor_entry.get().strip()  # 取得教授名稱關鍵字
    week_keyword = week_var.get().strip()

    selected_time = time_var.get().strip()
    time_index = time_options.index(selected_time)
    time_keyword = time_number[time_index]
    
    display_courses()

Label(Swindow, text="課程名稱：").place(x=550, y=10)
course_entry = Entry(Swindow, width=20)
course_entry.place(x=620, y=10)

Label(Swindow, text="課程代碼：").place(x=550, y=40)
number_entry = Entry(Swindow, width=20)
number_entry.place(x=620, y=40)

Label(Swindow, text="上課地點：").place(x=550, y=70)
room_entry = Entry(Swindow, width=20)
room_entry.place(x=620, y=70)

Label(Swindow, text="教授名稱：").place(x=550, y=100)
professor_entry = Entry(Swindow, width=20)
professor_entry.place(x=620, y=100)

Label(Swindow, text="星期：").place(x=780, y=15)
week_var = tk.StringVar(Swindow)
week_var.set(week_options[0])  # 設定初始值
week_menu = tk.OptionMenu(Swindow, week_var, *week_options)
week_menu.config(width=10)  # 設置下拉選單的寬度
week_menu.place(x=825, y=10)

Label(Swindow, text="時間：").place(x=780, y=45)
time_var = tk.StringVar(Swindow)
time_var.set(time_options[0])  # 設定初始值
time_menu = tk.OptionMenu(Swindow, time_var, *time_options)
time_menu.config(width=10)  # 設置下拉選單的寬度
time_menu.place(x=825, y=40)

Button(Swindow, text="搜尋", command=update_keywords).place(x=900, y=95)

def take_time(course_time):
    # 提取星期幾（例如：星期二）和時間區間（例如：13:00-15:00）
    _, time_range = course_time.split(" ")

    # 提取開始和結束時間（例如 "13:00-15:00"）
    start_time, end_time = time_range.split('-')

    # 提取並將開始時間和結束時間轉換為整數
    start_hour = int(start_time.split(":")[0])  # 提取 "13" 並轉為整數
    end_hour = int(end_time.split(":")[0])      # 提取 "15" 並轉為整數

    # 返回關鍵字詞及整數時間
    return start_hour, end_hour

def display_courses():
    # 先清空目前顯示的內容
    for widget in content_frame.winfo_children():
        widget.destroy()

    for col_idx, header in enumerate(headers):
        label = tk.Label(content_frame, text=header, borderwidth=1, relief="solid", padx=5, pady=5, bg="lightgray", width = 20)
        label.grid(row=0, column=col_idx, sticky="nsew", padx=2, pady=2)

    find = 0
        
    # 讀取並顯示課程列表
    for row_idx, row in enumerate(worksheet_courses.iter_rows(min_row=2, max_row=53, min_col=1, max_col=5, values_only=True), start=1):

        course_name = row[0]  # 假設課程名稱在第1列（索引0）
        code=row[1]
        number_name = str(row[1])
        week_name = row[2]
        start_hour, end_hour = take_time(row[2])
        room_name = row[3]
        professor_name = row[4]  # 假設教授名稱在第5列（索引4）

        if (course_keyword in course_name) and (number_keyword in number_name) and (room_keyword in room_name) and  (professor_keyword in professor_name) and  (week_keyword in week_name) and ((start_hour <= time_keyword and time_keyword < end_hour) or time_keyword == 0):

            find = 1

            # for col_idx, value in enumerate(row):
            #     label = tk.Label(data_frame, text=value if value else "", borderwidth=1, relief="solid", width=25, padx=4, pady=5)
            #     label.grid(row=row_idx + 1, column=col_idx, sticky="nsew", padx=2, pady=2)

            for col_idx, value in enumerate(row):
                button = tk.Button(content_frame, text=value if value else "", borderwidth=1, relief="solid", padx=4, pady=5, command=lambda code=code: show_course_details(code))
                button.grid(row=row_idx + 1, column=col_idx, sticky="nsew", padx=2, pady=2)

            action_frame = Frame(content_frame)
            action_frame.grid(row=row_idx + 1, column=len(headers) - 2, padx=2, pady=2, sticky="nsew")

            entry = tk.Entry(action_frame, width=10)
            entry.pack(side="left", padx=2, pady=2)

            def number_search_add(entry=entry, course_code=row[1]):
                number_search(entry, course_code, action="add")

            def number_search_drop(entry=entry, course_code=row[1]):
                number_search(entry, course_code, action="drop")

            button_add = Button(action_frame, text="加選", command=number_search_add)
            button_add.pack(side="left", padx=2, pady=2)

            button_drop = Button(action_frame, text="退選", command=number_search_drop)
            button_drop.pack(side="left", padx=2, pady=2)

            # 新增餘額確認按鈕
            def check_remaining_spots(course_code=row[1]):
                remaining_spots = get_course_remaining_spots(course_code)
                if remaining_spots is not None:
                    messagebox.showinfo("課程餘額", f"{all_course[course_code]['課程名稱']}{course_code}的剩餘名額為：{remaining_spots}")
                else:
                    messagebox.showinfo("課程餘額", f"{all_course[course_code]['課程名稱']}{course_code}的餘額資訊不可用")

            check_button = Button(content_frame, text="餘額確認", command=check_remaining_spots)
            check_button.grid(row=row_idx + 1, column=len(headers)-1, padx=2, pady=2, sticky="nsew")

            def number_search(entry=entry, course_code=row[1], action="add"):
                student_id = entry.get()
                if not student_id:
                    messagebox.showwarning("錯誤", "請輸入學號")
                    return

                entry.delete(0, END)
                path = search(student_id)

                if path:
                    total_credits = calculate_total_credits(path)
                    course_credit = get_course_credit(course_code)
                    if action == "add":
                        if samecourse_existing(path, course_code):
                            messagebox.showerror("重複課程", f"已加選相同課程：{all_course[course_code]['課程名稱']}{course_code}，加選失敗")
                            return
            
                        if is_course_in_schedule(path, course_code):
                            messagebox.showinfo("提醒", f"{all_course[course_code]['課程名稱']}{course_code}已存在於課表中")
                            return

                        if total_credits + course_credit > 25:
                            messagebox.showerror("超過學分上限", "加選失敗，超過學分上限！")
                            return

                        day, time_range = all_course[course_code]["開課時間"].split()
                        col_num, start_row, end_row = map_course_time_to_schedule(day, time_range)

                        if start_row and end_row:
                            if check_schedule_conflict(path, col_num, start_row, end_row):
                                messagebox.showwarning("衝堂", "衝堂，加選失敗")
                                return

                            remaining_spots = get_course_remaining_spots(course_code)
                            if remaining_spots is not None and remaining_spots > 0:
                                add_course_to_schedule(path, course_code, col_num, start_row, end_row)
                                update_course_remaining_spots(course_code, remaining_spots - 1)
                                #display_schedule(path)
                                messagebox.showinfo("成功", f"{all_course[course_code]['課程名稱']}{course_code}，加選成功")
                            else:
                                messagebox.showerror("加選失敗", "修課人數已滿，加選失敗")
                    elif action == "drop":
                        if not drop_course_from_schedule(path, course_code):
                            return
                        #display_schedule(path)

                else:
                    messagebox.showerror("錯誤", "找不到學號對應的課表")

    if(find == 0):
        # 顯示 "查無課程" 訊息
        not_found_label = tk.Label(content_frame, text="查無課程", fg="red", font=("Arial", 14, "bold"))
        not_found_label.grid(row=1, column=0, columnspan=len(headers), pady=10)  # 放在標題欄下方，跨越所有列
        print("not find")

display_courses()

Swindow.mainloop()