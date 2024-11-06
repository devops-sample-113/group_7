from tkinter import *
import tkinter as tk
import openpyxl
from tkinter import messagebox

Swindow = Tk()
Swindow.title("選課系統")
Swindow.geometry("1500x1000")  # 設置主視窗的大小

# 設置 Canvas 和 Scrollbar，將整個內容放入 Canvas 中
canvas = Canvas(Swindow, highlightthickness=0)
scrollbar = Scrollbar(Swindow, orient="vertical", command=canvas.yview)
scrollbar.pack(side="right", fill="y")
canvas.configure(yscrollcommand=scrollbar.set)
canvas.place(x=10, y=50, width=1480, height=700)

# 新增一個框架用於承載內容，並在 Canvas 中滾動
content_frame = Frame(canvas)
canvas.create_window((0, 0), window=content_frame, anchor="nw")

def on_configure(event):
    # 更新 Canvas 滾動範圍
    canvas.configure(scrollregion=canvas.bbox("all"))

content_frame.bind("<Configure>", on_configure)

# 綁定觸控面板/滑鼠滾輪滾動事件
def on_mouse_wheel(event):
    if event.delta > 0:  # 向上滾動
        canvas.yview_scroll(-1, "units")
    elif event.delta < 0:  # 向下滾動
        canvas.yview_scroll(1, "units")

canvas.bind_all("<MouseWheel>", on_mouse_wheel)  # Windows 和 Mac OS
canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))  # Linux 滾輪向上
canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))  # Linux 滾輪向下

def open_new_window():
    # 創建新的 Toplevel 視窗
    new_window = Toplevel(Swindow)
    new_window.title("課表頁面")
    new_window.geometry("400x300")  # 設置新視窗的大小

    # 在新視窗上添加一些元件
    label = Label(new_window, text="個人課表")
    label.pack(pady=20)

    close_button = Button(new_window, text="關閉", command=new_window.destroy)
    close_button.pack(pady=10)

# 按鍵，點擊後開啟新頁面
button_pop = Button(Swindow, text="課表頁面", command=open_new_window)
button_pop.place(x=10, y=10)

# 計算每列的寬度，使其隨 Canvas 的寬度自動調整
def adjust_column_widths(event=None):
    canvas_width = canvas.winfo_width()
    col_width = canvas_width // len(headers)  # 每列寬度
    for col, header in enumerate(headers):
        label = tk.Label(content_frame, text=header, font=("Arial", 10, "bold"), borderwidth=1, relief="solid", width=col_width//10)
        label.grid(row=0, column=col, sticky="nsew", padx=2, pady=2)

# 更新標題行和資料行的寬度
canvas.bind("<Configure>", adjust_column_widths)

# 標題行
headers = ["課程名稱", "課程代碼", "開課時間", "上課地點", "授課教授", "加退選匡"]

# 開啟 Excel 文件
path = '資料庫.xlsx'
workbook = openpyxl.load_workbook(path)
worksheet_courses = workbook["課程"]
worksheet_students = workbook["學生"]

def search(id):
    for row in worksheet_students.iter_rows(min_row=2, values_only=True):
        name, id_, schedule_path = row[:3]
        if id_ == id:
            schedule_path = f"個人課表/{id}.xlsx"
            return schedule_path
    return None

# Function to check if a course is in the student's schedule
def is_course_in_schedule(schedule_path, course_code):
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[1] == course_code:  # Assuming course codes are in the second column of the schedule
            return True
    return False

def map_course_time_to_schedule(day, time_range):
    # Map day to row, e.g., "星期一" to row 2, "星期二" to row 3, etc.
    day_mapping = {"星期一": 2, "星期二": 3, "星期三": 4, "星期四": 5, "星期五": 6}
    start_time, end_time = time_range.split('-')
    col_num = day_mapping[day]

    # Map time to columns: assuming 14:00-16:00 corresponds to columns 8 and 9
    if start_time == "08:00":
        start_row, end_row = 2, 3  # For example: columns 8 and 9 for 08:00-10:00
    elif start_time == "09:00":
        start_row, end_row = 3, 4
    elif start_time == "10:00":
        start_row, end_row = 4, 5
    elif start_time == "11:00":
        start_row, end_row = 5, 6
    elif start_time == "12:00":
        start_row, end_row = 6, 7
    elif start_time == "13:00":
        start_row, end_row = 7, 8
    elif start_time == "14:00":
        start_row, end_row = 8, 9
    elif start_time == "15:00":
        start_row, end_row = 9, 10
    # Add more mappings for different times if needed
    else:
        raise ValueError(f"Unrecognized time slot: {start_time}-{end_time}")

    return col_num, start_row, end_row

def add_course_to_schedule(schedule_path, course_name, col_num, start_row, end_row):
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active
    # Check for duplicate course name in the schedule
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
        if course_name in row:
            # If the course name is found, show an error message and exit
            messagebox.showerror("重複課程", "已加選其他相同課程，加選失敗")
            return
    # Check for conflict: see if cells in the specified range already contain a course
    if worksheet.cell(row=start_row, column=col_num).value or worksheet.cell(row=end_row, column=col_num).value:
        # If either cell is occupied, show an error message and return
        messagebox.showwarning("衝堂", "衝堂，加選失敗")
        return
    # Write the course name in specified columns
    worksheet.cell(row=start_row, column=col_num).value = course_name
    worksheet.cell(row=end_row, column=col_num).value = course_name
    workbook.save(schedule_path)

# 讀取 1-39 行，A-E 列的內容顯示在 Label 中，並在「加退選匡」列新增輸入框和確認按鈕
for row_idx, row in enumerate(worksheet_courses.iter_rows(min_row=2, max_row=53, min_col=1, max_col=5, values_only=True), start=1):
    for col_idx, value in enumerate(row):
        label = tk.Label(content_frame, text=value if value is not None else "", borderwidth=1, relief="solid", padx=5, pady=5)
        label.grid(row=row_idx, column=col_idx, sticky="nsew", padx=2, pady=2)
    
    # 在「加退選匡」列添加輸入框和確認按鈕
    entry = tk.Entry(content_frame, width=15)
    entry.grid(row=row_idx, column=len(headers) - 1, padx=2, pady=2, sticky="nsew")
    
    # 確認按鈕功能
    def number_search(entry=entry, course_code=row[0], course_name=row[0], course_time=row[2]):
        # 取得輸入資料
        student_id = entry.get()
        if not student_id:
            messagebox.showwarning("錯誤", "請輸入學號")  # 顯示提示框
        else:
            entry.delete(0, END)  # 清空輸入框
            path = search(student_id)
            # Debugging: 打印學號和路徑來檢查正確性
            #print(f"輸入的學號: {student_id}, 生成的課表路徑: {path}")
            if path:
                # 若找到課表路徑，則開啟新視窗並顯示課表
                #messagebox.showinfo("成功", f"找到課表：{path}")
                if is_course_in_schedule(path, course_code):
                    messagebox.showinfo("提醒", f"課程 {course_code} 已存在於課表中")
                else:
                    # Parse the course time to find the correct row and columns in the schedule
                    day, time_range = course_time.split()
                    col_num, start_row, end_row = map_course_time_to_schedule(day, time_range)
                
                    # Write the course name in the corresponding slots
                    add_course_to_schedule(path, course_name, col_num, start_row, end_row)

                    #messagebox.showinfo("訊息", f"課程 {course_code} 已成功加入至課表")
                    #messagebox.showinfo("訊息", start_row)
                display_schedule(path)
            else:
                messagebox.showinfo("錯誤", "學號輸入錯誤")

    def display_schedule(schedule_path):
        # 開啟新視窗顯示課表
        schedule_window = Toplevel(Swindow)
        schedule_window.title("個人課表")
        schedule_window.geometry("600x400")

        workbook = openpyxl.load_workbook(schedule_path)
        worksheet = workbook.active

        # 顯示課表內容
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row= 10, values_only=True)):
            for col_idx, value in enumerate(row):
                label = tk.Label(schedule_window, text=value if value else "", borderwidth=1, relief="solid", padx=5, pady=5)
                label.grid(row=row_idx, column=col_idx, sticky="nsew", padx=2, pady=2)

        # 關閉按鈕
        close_button = Button(schedule_window, text="關閉", command=schedule_window.destroy)
        close_button.grid(row=row_idx + 1, column=0, columnspan=len(row), pady=10)

    confirm_button = tk.Button(content_frame, text="確認", command=number_search)
    confirm_button.grid(row=row_idx, column=len(headers), padx=2, pady=2, sticky="nsew")

Swindow.mainloop()