import openpyxl

# 設定課表檔案路徑
course_file_path = "C:\\Users\\User\\OneDrive - 逢甲大學\\文件\\GitHub\\group_7\\個人課表\\D0987654.xlsx"
course_sheet = openpyxl.load_workbook(course_file_path)["課表"]

# 設定可修課餘額檔案路徑
balance_file_path = "C:\\Users\\User\\OneDrive - 逢甲大學\\文件\\GitHub\\group_7\\可修課餘額.xlsx"
balance_sheet = openpyxl.load_workbook(balance_file_path)["Sheet1"]  # 假設工作表名稱為 Sheet1

def remove_all_courses(sheet, course_name):
    found = False  # 用於追蹤是否有找到該課程
    # 假設課表是從 A1 開始的 9 行 9 列範圍
    for row in sheet.iter_rows(min_row=1, max_row=9, min_col=1, max_col=9):
        for cell in row:
            if cell.value == course_name:
                cell.value = None  # 將儲存格的值設為 None（移除內容）
                found = True  # 標記已找到課程
    return found  # 返回是否有移除任何內容

def update_balance(sheet, row, col):
    current_balance = sheet.cell(row=row, column=col).value  # 讀取當前可修課餘額
    if current_balance is None:
        current_balance = 0  # 如果儲存格是空的，則設定為 0
    sheet.cell(row=row, column=col, value=current_balance + 1)  # 將餘額增加 1

# 示例：移除所有出現的課程名稱
course_name_to_remove = "軟體測試"  # 您要移除的課程名稱

if remove_all_courses(course_sheet, course_name_to_remove):
    print(f"課程 {course_name_to_remove} 已移除")
    # 更新可修課餘額，假設在第 1 行第 1 列
    update_balance(balance_sheet, 1, 1)
    print("可修課餘額已增加 1")
else:
    print(f"課程 {course_name_to_remove} 未找到")

# 儲存變更到 Excel 檔案
course_sheet.parent.save(course_file_path)
balance_sheet.parent.save(balance_file_path)
