from tkinter import *
import tkinter as tk
import openpyxl
from tkinter import messagebox

Swindow = Tk()
Swindow.title("選課系統")
Swindow.geometry("1500x1000")  # 設置主視窗的大小

# 設置 Canvas 和 Scrollbar，將整個內容放入 Canvas 中
canvas = Canvas(Swindow, highlightthickness=0)
scrollbar = Scrollbar(Swindow, orient="vertical", command=canvas.yview)
scrollbar.pack(side="right", fill="y")
canvas.configure(yscrollcommand=scrollbar.set)
canvas.place(x=10, y=130, width=1480, height=700)

# 新增一個框架用於承載內容，並在 Canvas 中滾動
content_frame = Frame(canvas)
canvas.create_window((0, 0), window=content_frame, anchor="nw")

# 用於顯示課程資料的子框架    ##查
data_frame = Frame(content_frame)
data_frame.grid(row=1, column=0, columnspan=6, sticky="nsew")

def on_configure(event):
    # 更新 Canvas 滾動範圍
    canvas.configure(scrollregion=canvas.bbox("all"))

content_frame.bind("<Configure>", on_configure)

# 綁定觸控面板/滑鼠滾輪滾動事件
def on_mouse_wheel(event):
    if event.delta > 0:  # 向上滾動
        canvas.yview_scroll(-1, "units")
    elif event.delta < 0:  # 向下滾動
        canvas.yview_scroll(1, "units")

canvas.bind_all("<MouseWheel>", on_mouse_wheel)  # Windows 和 Mac OS
canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))  # Linux 滾輪向上
canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))  # Linux 滾輪向下

def open_new_window():
    # 創建新的 Toplevel 視窗
    new_window = Toplevel(Swindow)
    new_window.title("課表頁面")
    new_window.geometry("400x300")  # 設置新視窗的大小

    # 在新視窗上添加一些元件
    label = Label(new_window, text="個人課表")
    label.pack(pady=20)

    close_button = Button(new_window, text="關閉", command=new_window.destroy)
    close_button.pack(pady=10)

# 按鍵，點擊後開啟新頁面
button_pop = Button(Swindow, text="課表頁面", command=open_new_window)
button_pop.place(x=10, y=10)

# 計算每列的寬度，使其隨 Canvas 的寬度自動調整
def adjust_column_widths(event=None):
    canvas_width = canvas.winfo_width()
    col_width = canvas_width // len(headers)  # 每列寬度
    for col, header in enumerate(headers):
        label = tk.Label(content_frame, text=header, font=("Arial", 10, "bold"), borderwidth=1, relief="solid", width=col_width//10)
        label.grid(row=0, column=col, sticky="nsew", padx=2, pady=2)

# 更新標題行和資料行的寬度
canvas.bind("<Configure>", adjust_column_widths)

# 標題行
headers = ["課程名稱", "課程代碼", "開課時間", "上課地點", "授課教授", "加退選匡"]

# 設定標題行顯示
for col, header in enumerate(headers):
    label = tk.Label(content_frame, text=header, font=("Arial", 10, "bold"), borderwidth=1, relief="solid", width=15)
    label.grid(row=0, column=col, sticky="nsew", padx=2, pady=2)

# 開啟 Excel 文件
path = '資料庫.xlsx'
workbook = openpyxl.load_workbook(path)
worksheet_courses = workbook["課程"]
worksheet_students = workbook["學生"]

def search(id):
    for row in worksheet_students.iter_rows(min_row=2, values_only=True):
        name, id_, schedule_path = row[:3]
        if id_ == id:
            schedule_path = f"/個人課表/{id}.xlsx"
            return schedule_path
    return None

###########################################搜尋

course_keyword = ""
number_keyword = ""
room_keyword = ""
professor_keyword = ""
week_keyword = ""
time_keyword = 0
week_options = ["", "一", "二", "三", "四", "五"]
time_options = ["", "08:00-09:00", "09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00", "16:00-17:00"]
time_number = [0, 8, 9, 10, 11, 12, 13, 14, 15, 16]

def update_keywords():
    global professor_keyword, course_keyword, number_keyword, room_keyword, week_keyword, time_keyword  # 讓這兩個變數可以在函數外部使用並更新
    course_keyword = course_entry.get().strip()
    number_keyword = number_entry.get().strip()
    room_keyword = room_entry.get().strip()
    professor_keyword = professor_entry.get().strip()  # 取得教授名稱關鍵字
    week_keyword = week_var.get().strip()

    selected_time = time_var.get().strip()
    time_index = time_options.index(selected_time)
    time_keyword = time_number[time_index]
    
    display_courses()

Label(Swindow, text="課程名稱：").place(x=510, y=10)
course_entry = Entry(Swindow, width=20)
course_entry.place(x=600, y=10)

Label(Swindow, text="課程代碼：").place(x=510, y=40)
number_entry = Entry(Swindow, width=20)
number_entry.place(x=600, y=40)

Label(Swindow, text="上課地點：").place(x=510, y=70)
room_entry = Entry(Swindow, width=20)
room_entry.place(x=600, y=70)

Label(Swindow, text="教授名稱：").place(x=510, y=100)
professor_entry = Entry(Swindow, width=20)
professor_entry.place(x=600, y=100)

Label(Swindow, text="星期：").place(x=780, y=10)
week_var = tk.StringVar(Swindow)
week_var.set(week_options[0])  # 設定初始值
week_menu = tk.OptionMenu(Swindow, week_var, *week_options)
week_menu.config(width=10)  # 設置下拉選單的寬度
week_menu.place(x=870, y=10)

Label(Swindow, text="時間：").place(x=780, y=40)
time_var = tk.StringVar(Swindow)
time_var.set(time_options[0])  # 設定初始值
time_menu = tk.OptionMenu(Swindow, time_var, *time_options)
time_menu.config(width=10)  # 設置下拉選單的寬度
time_menu.place(x=870, y=40)

Button(Swindow, text="搜尋", command=update_keywords).place(x=1050, y=25)

def take_time(course_time):
    # 提取星期幾（例如：星期二）和時間區間（例如：13:00-15:00）
    _, time_range = course_time.split(" ")

    # 提取開始和結束時間（例如 "13:00-15:00"）
    start_time, end_time = time_range.split('-')

    # 提取並將開始時間和結束時間轉換為整數
    start_hour = int(start_time.split(":")[0])  # 提取 "13" 並轉為整數
    end_hour = int(end_time.split(":")[0])      # 提取 "15" 並轉為整數

    # 返回關鍵字詞及整數時間
    return start_hour, end_hour

def display_courses():
    # 先清空目前顯示的內容
    for widget in data_frame.winfo_children():
        widget.destroy()

    find = 0

    # 讀取 1-39 行，A-E 列的內容顯示在 Label 中，並在「加退選匡」列新增輸入框和確認按鈕
    for row_idx, row in enumerate(worksheet_courses.iter_rows(min_row=2, max_row=53, min_col=1, max_col=5, values_only=True), start=1):
        
        course_name = row[0]  # 假設課程名稱在第1列（索引0）
        number_name = str(row[1])
        week_name = row[2]
        start_hour, end_hour = take_time(row[2])
        room_name = row[3]
        professor_name = row[4]  # 假設教授名稱在第5列（索引4）
        
        if (course_keyword in course_name) and (number_keyword in number_name) and (room_keyword in room_name) and  (professor_keyword in professor_name) and  (week_keyword in week_name) and ((start_hour <= time_keyword and time_keyword < end_hour) or time_keyword == 0):

            find = 1
        # 顯示符合條件的課程
            for col_idx, value in enumerate(row):
                label = tk.Label(data_frame, text=value if value is not None else "", borderwidth=1, relief="solid", padx=5, pady=5)
                label.grid(row=row_idx, column=col_idx, sticky="nsew", padx=2, pady=2)
            
            # 在「加退選匡」列添加輸入框和確認按鈕
            entry = tk.Entry(data_frame, width=15)
            entry.grid(row=row_idx, column=len(headers) - 1, padx=2, pady=2, sticky="nsew")
            
            # 確認按鈕功能
            def number_search(entry=entry):
                # 取得輸入資料
                student_id = entry.get()
                if not student_id:
                    messagebox.showwarning("錯誤", "請輸入學號")  # 顯示提示框
                else:
                    # 儲存學號並清空輸入框
                    with open("data.txt", "w") as file:
                        file.write(student_id)
                    entry.delete(0, END)  # 清空輸入框
                    ##獲取輸入資料(學號)
                    try:
                        with open("data.txt","r") as file:
                            number=file.read()
                    except FileNotFoundError:
                        number="輸入錯誤"
                    
                    path = search(number)
                    messagebox.showinfo("正確", path)
                    if path:
                        # 若找到課表路徑，則開啟新視窗並顯示課表
                        display_schedule(path)
                    else:
                        messagebox.showinfo("錯誤", "學號輸入錯誤")
                    ##將data.txt暫存資料清空
                    with open("data.txt","w") as file:
                        file.write("")

            def display_schedule(schedule_path):
                # 開啟新視窗顯示課表
                schedule_window = Toplevel(Swindow)
                schedule_window.title("個人課表")
                schedule_window.geometry("600x400")

                workbook = openpyxl.load_workbook(schedule_path)
                worksheet = workbook.active

                # 顯示課表內容
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, values_only=True)):
                    for col_idx, value in enumerate(row):
                        label = tk.Label(schedule_window, text=value if value else "", borderwidth=1, relief="solid", padx=5, pady=5)
                        label.grid(row=row_idx, column=col_idx, sticky="nsew", padx=2, pady=2)

                # 關閉按鈕
                close_button = Button(schedule_window, text="關閉", command=schedule_window.destroy)
                close_button.grid(row=row_idx + 1, column=0, columnspan=len(row), pady=10)

            confirm_button = tk.Button(data_frame, text="確認", command=number_search)
            confirm_button.grid(row=row_idx, column=len(headers), padx=2, pady=2, sticky="nsew")

    if(find == 0):
        # 顯示 "查無課程" 訊息
        not_found_label = tk.Label(data_frame, text="查無課程", fg="red", font=("Arial", 14, "bold"))
        not_found_label.grid(row=1, column=0, columnspan=len(headers), pady=10)  # 放在標題欄下方，跨越所有列
        print("not find")

display_courses()

Swindow.mainloop()