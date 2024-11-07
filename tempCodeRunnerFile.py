from tkinter import *
import tkinter as tk
import openpyxl
from tkinter import messagebox

Swindow = Tk()
Swindow.title("選課系統")
Swindow.geometry("1500x1000")

# 設置 Canvas 和 Scrollbar
canvas = Canvas(Swindow, highlightthickness=0)
scrollbar = Scrollbar(Swindow, orient="vertical", command=canvas.yview)
scrollbar.pack(side="right", fill="y")
canvas.configure(yscrollcommand=scrollbar.set)
canvas.place(x=10, y=50, width=1480, height=700)

# 新增框架用於承載內容
content_frame = Frame(canvas)
canvas.create_window((0, 0), window=content_frame, anchor="nw")

def on_configure(event):
    canvas.configure(scrollregion=canvas.bbox("all"))

content_frame.bind("<Configure>", on_configure)

def on_mouse_wheel(event):
    canvas.yview_scroll(-1 if event.delta > 0 else 1, "units")

canvas.bind_all("<MouseWheel>", on_mouse_wheel)  # Windows 和 Mac OS
#canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))  # Linux 向上
#canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))  # Linux 向下

# 開啟課表頁面
def open_new_window():
    new_window = Toplevel(Swindow)
    new_window.title("課表頁面")
    new_window.geometry("400x300")

    Label(new_window, text="個人課表").pack(pady=20)
    Button(new_window, text="關閉", command=new_window.destroy).pack(pady=10)

button_pop = Button(Swindow, text="課表頁面", command=open_new_window)
button_pop.place(x=10, y=10)

headers = ["課程名稱", "課程代碼", "開課時間", "上課地點", "授課教授", "加退選匡"]
# 在 content_frame 中加入標題行
for col_idx, header in enumerate(headers):
    label = tk.Label(content_frame, text=header, borderwidth=1, relief="solid", padx=5, pady=5, bg="lightgray")
    label.grid(row=0, column=col_idx, sticky="nsew", padx=2, pady=2)

# 開啟 Excel 文件
path = '資料庫.xlsx'
workbook = openpyxl.load_workbook(path)
worksheet_courses = workbook["課程"]
worksheet_students = workbook["學生"]

# 初始化 all_course 字典
all_course = {}
for row in worksheet_courses.iter_rows(min_row=2, values_only=True):
    course_name, course_code, course_time, location, professor, remaining_spots, credit = row[:7]
    all_course[course_code] = {
        "課程名稱": course_name,
        "開課時間": course_time,
        "上課地點": location,
        "授課教授": professor,
        "目前可修課人數餘額": remaining_spots,
        "學分": credit
    }

# 查詢學號
def search(id):
    for row in worksheet_students.iter_rows(min_row=2, values_only=True):
        name, id_, schedule_path = row[:3]
        if id_ == id:
            return f"個人課表/{id}.xlsx"
    return None

# 取得課程學分
def get_course_credit(course_code):
    return all_course.get(course_code, {}).get("學分")

# 計算總學分
def calculate_total_credits(schedule_path):
    total_credits = 0
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active
    added_courses = set()

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if cell and cell not in added_courses:
                course_credit = get_course_credit(cell)
                if course_credit:
                    total_credits += course_credit
                    added_courses.add(cell)
    return total_credits

#檢查是否已含有相同名稱課程
def samecourse_existing(schedule_path, course_code):
    course_name = all_course[course_code]["課程名稱"]
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
        for cell in row:
            if cell and cell != course_code:  # 確保排除當前課程代碼
                # 取得該課程代碼的名稱
                existing_course_name = all_course.get(cell, {}).get('課程名稱')
                if existing_course_name == course_name:
                    return True
    return False

# 檢查課程是否在學生課表中
def is_course_in_schedule(schedule_path, course_code):
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if cell == course_code:
                return True
    return False

# 將課程時間對應至課表
def map_course_time_to_schedule(day, time_range):
    day_mapping = {"星期一": 2, "星期二": 3, "星期三": 4, "星期四": 5, "星期五": 6}
    start_time, _ = time_range.split('-')
    col_num = day_mapping[day]

    time_mapping = {
        "08:00": (2, 3), "09:00": (3, 4), "10:00": (4, 5),
        "11:00": (5, 6), "12:00": (6, 7), "13:00": (7, 8),
        "14:00": (8, 9), "15:00": (9, 10)
    }
    return col_num, *time_mapping.get(start_time, (None, None))

# 檢查課程是否衝堂
def check_schedule_conflict(schedule_path, col_num, start_row, end_row):
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active

    # 檢查開始與結束行的位置是否已經有課程
    if worksheet.cell(row=start_row, column=col_num).value or worksheet.cell(row=end_row, column=col_num).value:
        return True
    return False

# 獲取課程剩餘名額
def get_course_remaining_spots(course_code):
    return all_course.get(course_code, {}).get("目前可修課人數餘額")

# 新增課程至課表
def add_course_to_schedule(schedule_path, course_code, col_num, start_row, end_row):
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active
    
    worksheet.cell(row=start_row, column=col_num).value = course_code
    worksheet.cell(row=end_row, column=col_num).value = course_code
    workbook.save(schedule_path)

# 更新課程剩餘名額
def update_course_remaining_spots(course_code, new_remaining_spots):
    if course_code in all_course:
        all_course[course_code]["目前可修課人數餘額"] = new_remaining_spots
        for row in worksheet_courses.iter_rows(min_row=2, values_only=False):
            if row[1].value == course_code:
                row[13].value = new_remaining_spots
                workbook.save(path)
                break

# 顯示課表
def display_schedule(schedule_path):
    schedule_window = Toplevel(Swindow)
    schedule_window.title("個人課表")
    schedule_window.geometry("600x400")

    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True)):
        for col_idx, value in enumerate(row):
            label = tk.Label(schedule_window, text=value if value else "", borderwidth=1, relief="solid", padx=5, pady=5)
            label.grid(row=row_idx, column=col_idx, sticky="nsew", padx=2, pady=2)

    Button(schedule_window, text="關閉", command=schedule_window.destroy).grid(row=row_idx + 1, column=0, columnspan=len(row), pady=10)

# 退選課程
def drop_course_from_schedule(schedule_path, course_code):
    # 讀取課表
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active

    found = False
    for row in worksheet.iter_rows(min_row=2, values_only=False):
        for cell in row:
            if cell.value == course_code:
                cell.value = None  # 刪除該課程
                found = True
                break

    if not found:
        messagebox.showinfo("退選失敗", f"課表中並沒有 {all_course[course_code]['課程名稱']}")
        return False

    # 退選後檢查學分
    total_credits = calculate_total_credits(schedule_path)
    course_credit = get_course_credit(course_code)
    if total_credits - course_credit < 9:
        messagebox.showerror("低於學分下限", "退選失敗，學分低於 9 學分！")
        return False

    workbook.save(schedule_path)
    messagebox.showinfo("退選成功", f"{all_course[course_code]['課程名稱']} 退選成功")
    return True


# 讀取並顯示課程列表
for row_idx, row in enumerate(worksheet_courses.iter_rows(min_row=2, max_row=53, min_col=1, max_col=5, values_only=True), start=1):
    for col_idx, value in enumerate(row):
        label = tk.Label(content_frame, text=value if value else "", borderwidth=1, relief="solid", padx=5, pady=5)
        label.grid(row=row_idx + 1, column=col_idx, sticky="nsew", padx=2, pady=2)
    
    entry = tk.Entry(content_frame, width=15)
    entry.grid(row=row_idx + 1, column=len(headers) - 1, padx=2, pady=2, sticky="nsew")
    def number_search_add(entry=entry, course_code=row[1]):
        number_search(entry, course_code, action="add")

    def number_search_drop(entry=entry, course_code=row[1]):
        number_search(entry, course_code, action="drop")
        
    button_add = Button(content_frame, text="加選", command=number_search_add)
    button_add.grid(row=row_idx + 1, column=len(headers), padx=2, pady=2, sticky="nsew")

    button_drop = Button(content_frame, text="退選", command=number_search_drop)
    button_drop.grid(row=row_idx + 1, column=len(headers) + 1, padx=2, pady=2, sticky="nsew")

    def number_search(entry=entry, course_code=row[1], action="add"):
        student_id = entry.get()
        if not student_id:
            messagebox.showwarning("錯誤", "請輸入學號")
            return

        entry.delete(0, END)
        path = search(student_id)

        if path:
            total_credits = calculate_total_credits(path)
            course_credit = get_course_credit(course_code)
            if action == "add":
                if samecourse_existing(path, course_code):
                    messagebox.showerror("重複課程", f"課表中已有相同課程：{all_course[course_code]['課程名稱']}，加選失敗")
                    return
    
                if is_course_in_schedule(path, course_code):
                    messagebox.showinfo("提醒", f"{all_course[course_code]['課程名稱']} 已存在於課表中")
                    return

                if total_credits + course_credit > 25:
                    messagebox.showerror("超過學分上限", "加選失敗，超過學分上限！")
                    return

                day, time_range = all_course[course_code]["開課時間"].split()
                col_num, start_row, end_row = map_course_time_to_schedule(day, time_range)

                if start_row and end_row:
                    if check_schedule_conflict(path, col_num, start_row, end_row):
                        messagebox.showwarning("衝堂", "衝堂，加選失敗")
                        return

                    remaining_spots = get_course_remaining_spots(course_code)
                    if remaining_spots is not None and remaining_spots > 0:
                        add_course_to_schedule(path, course_code, col_num, start_row, end_row)
                        update_course_remaining_spots(course_code, remaining_spots - 1)
                        display_schedule(path)
                        messagebox.showinfo("成功", f"{all_course[course_code]['課程名稱']} 加選成功")
                    else:
                        messagebox.showerror("加選失敗", "該課程無剩餘名額")
            elif action == "drop":
                if not drop_course_from_schedule(path, course_code):
                    return
                display_schedule(path)

        else:
            messagebox.showerror("錯誤", "找不到學號對應的課表")

Swindow.mainloop()