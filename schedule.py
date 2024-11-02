from tkinter import *
import tkinter as tk
import openpyxl
import subprocess

##返回課表搜尋頁面
def return_search():
    Swindow.destroy()  # 關閉當前視窗
    subprocess.Popen(["python", "search_schedule.py"])  # 執行第二個程式

##課表頁面視窗初始設定
Swindow=Tk()
Swindow.title("你的個人課表")
Swindow.geometry('540x430+390+75')

##獲取輸入資料(學號)
try:
    with open("data.txt","r") as file:
        number=file.read()
except FileNotFoundError:
    number="輸入錯誤"


##搜尋學號/教師證號/助教證號對應的個人課表路徑
def search(id):
    #開啟excel
    path='C:\\Users\\User\\Documents\\GitHub\\group_7\\資料庫.xlsx'
    workbook=openpyxl.load_workbook(path)

    #選擇工作表
    if number[0]=="D":
        worksheet=workbook["學生"]
    elif number[0]=="T":
        worksheet=workbook["教授"]
    elif number[0]=="A":
        worksheet=workbook["助教"]
    else: return "無效的格式"

    for row in worksheet.iter_rows(min_row=2, values_only=True):  # 從第2行開始（假設第1行是標題）
        name, id_, schedule_path = row[:3]  # 解包每一行的資料

        # 如果學號符合，回傳個人課表路徑
        if id_ == id:
            return schedule_path
        
    return f"找不到 {id} "

#查詢課表
path=search(number)

if "找不到" not in path and "無效的學號" not in path:
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active

    #判斷課表是否為空
    is_empty=all(all(cell is None for cell in row) for row in worksheet.iter_rows(min_row=2, max_row=10, min_col=2, max_col=6, values_only=True))

    if is_empty:
        ##顯示下方提示文字-沒有課程
        label_prompt = tk.Label(Swindow, text="尚未加選任何課程", fg="blue",font=20)
        label_prompt.place(x=270,y=190,anchor="center")

    else:

        ##個人課表顯示
        # 讀取 1-10 行，A-F 列的內容並顯示在 Label 中
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=6, values_only=True)):
            for col_idx, value in enumerate(row):
                label = tk.Label(Swindow, text=value if value is not None else "", borderwidth=1, relief="solid", padx=5, pady=5)
                label.grid(row=row_idx, column=col_idx, sticky="nsew", padx=2, pady=2)

        ## 設定固定列寬和行高
        column_width = 80
        row_height = 30

        # 設定每一列的寬度
        for col in range(6):
            Swindow.grid_columnconfigure(col, minsize=column_width)  # 設定最小列寬

        # 設定每一行的高度
        for row in range(10):  # 假設有10行
            Swindow.grid_rowconfigure(row, minsize=row_height)  # 設定最小行高

else:
    label_error = tk.Label(Swindow, text=path, fg="red",font=20)
    label_error.place(x=270,y=190,anchor="center")

##顯示下方提示文字-查詢證號
label_prompt = tk.Label(Swindow, text="          正在查詢：")
label_prompt.place(x=50,y=360,anchor="center")

label_data = tk.Label(Swindow, text=number)
label_data.place(x=150,y=360,anchor="center")

##將data.txt暫存資料清空
with open("data.txt","w") as file:
        file.write("")

##返回按鈕
back=Button(Swindow,text="返回",anchor="s",command=return_search)
back.place(x=500,y=360,anchor="center")

Swindow.mainloop()