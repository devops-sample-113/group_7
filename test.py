from tkinter import *
import tkinter as tk
import openpyxl
from tkinter import messagebox

Swindow = Tk()
Swindow.title("選課系統")
Swindow.geometry("1500x1000")

# 設置 Canvas 和 Scrollbar
canvas = Canvas(Swindow, highlightthickness=0)
scrollbar = Scrollbar(Swindow, orient="vertical", command=canvas.yview)
scrollbar.pack(side="right", fill="y")
canvas.configure(yscrollcommand=scrollbar.set)
canvas.place(x=10, y=50, width=1480, height=700)

# 新增框架用於承載內容
content_frame = Frame(canvas)
canvas.create_window((0, 0), window=content_frame, anchor="nw")

# 更新 Canvas 滾動範圍
def on_configure(event):
    canvas.configure(scrollregion=canvas.bbox("all"))

content_frame.bind("<Configure>", on_configure)

# 綁定滾輪滾動事件
def on_mouse_wheel(event):
    canvas.yview_scroll(-1 if event.delta > 0 else 1, "units")

canvas.bind_all("<MouseWheel>", on_mouse_wheel)  # Windows 和 Mac OS
canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))  # Linux 向上
canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))  # Linux 向下

# 開啟課表頁面
def open_new_window():
    new_window = Toplevel(Swindow)
    new_window.title("課表頁面")
    new_window.geometry("400x300")

    Label(new_window, text="個人課表").pack(pady=20)
    Button(new_window, text="關閉", command=new_window.destroy).pack(pady=10)

button_pop = Button(Swindow, text="課表頁面", command=open_new_window)
button_pop.place(x=10, y=10)

# 計算每列的寬度，使其隨 Canvas 的寬度自動調整
def adjust_column_widths(event=None):
    canvas_width = canvas.winfo_width()
    col_width = canvas_width // len(headers)  # 每列寬度
    for col, header in enumerate(headers):
        label = tk.Label(content_frame, text=header, font=("Arial", 10, "bold"), borderwidth=1, relief="solid", width=col_width//10)
        label.grid(row=0, column=col, sticky="nsew", padx=2, pady=2)

# 更新標題行和資料行的寬度
canvas.bind("<Configure>", adjust_column_widths)

# 設置標題行
headers = ["課程名稱", "課程代碼", "開課時間", "上課地點", "授課教授", "加退選匡"]

# 開啟 Excel 文件
path = '資料庫.xlsx'
workbook = openpyxl.load_workbook(path)
worksheet_courses = workbook["課程"]
worksheet_students = workbook["學生"]

# 查詢學號
def search(id):
    for row in worksheet_students.iter_rows(min_row=2, values_only=True):
        name, id_, schedule_path = row[:3]
        if id_ == id:
            return f"個人課表/{id}.xlsx"
    return None

# 檢查課程是否在學生課表中
def is_course_in_schedule(schedule_path, course_name):
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[1] == course_name:
            return True
    return False

# 將課程時間對應至課表
def map_course_time_to_schedule(day, time_range):
    day_mapping = {"星期一": 2, "星期二": 3, "星期三": 4, "星期四": 5, "星期五": 6}
    start_time, _ = time_range.split('-')
    col_num = day_mapping[day]

    time_mapping = {
        "08:00": (2, 3), "09:00": (3, 4), "10:00": (4, 5),
        "11:00": (5, 6), "12:00": (6, 7), "13:00": (7, 8),
        "14:00": (8, 9), "15:00": (9, 10)
    }
    return col_num, *time_mapping.get(start_time, (None, None))

# 取得課程學分
def get_course_credit(course_name):
    for row in worksheet_courses.iter_rows(min_row=2, values_only=True):
        if row[0] == course_name:
            return row[14]
    return None

# 計算總學分
def calculate_total_credits(schedule_path):
    total_credits = 0
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active
    added_courses = set()

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if cell and cell not in added_courses:
                course_credit = get_course_credit(cell)
                if course_credit:
                    total_credits += course_credit
                    added_courses.add(cell)
    return total_credits

# 獲取課程剩餘名額
def get_course_remaining_spots(course_name):
    for row in worksheet_courses.iter_rows(min_row=2, values_only=True):
        if row[0] == course_name:
            return row[13]  # 剩餘名額在第14欄（index 13）
    return None

# 更新課程剩餘名額
def update_course_remaining_spots(course_name, new_remaining_spots):
    for row in worksheet_courses.iter_rows(min_row=2, values_only=False):
        if row[0].value == course_name:
            row[13].value = new_remaining_spots  # 更新第14欄（index 13）的值
            workbook.save(path)
            break

# 新增課程至課表
def add_course_to_schedule(schedule_path, course_name, course_code, col_num, start_row, end_row):
    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
        if course_name in row:
            messagebox.showerror("重複課程", "已加選相同課程，加選失敗")
            return

    if worksheet.cell(row=start_row, column=col_num).value or worksheet.cell(row=end_row, column=col_num).value:
        messagebox.showwarning("衝堂", "衝堂，加選失敗")
        return

    worksheet.cell(row=start_row, column=col_num).value = course_name
    worksheet.cell(row=end_row, column=col_num).value = course_name
    workbook.save(schedule_path)

# 顯示課表
def display_schedule(schedule_path):
    schedule_window = Toplevel(Swindow)
    schedule_window.title("個人課表")
    schedule_window.geometry("600x400")

    workbook = openpyxl.load_workbook(schedule_path)
    worksheet = workbook.active

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True)):
        for col_idx, value in enumerate(row):
            label = tk.Label(schedule_window, text=value if value else "", borderwidth=1, relief="solid", padx=5, pady=5)
            label.grid(row=row_idx, column=col_idx, sticky="nsew", padx=2, pady=2)

    Button(schedule_window, text="關閉", command=schedule_window.destroy).grid(row=row_idx + 1, column=0, columnspan=len(row), pady=10)

# 讀取並顯示課程列表
for row_idx, row in enumerate(worksheet_courses.iter_rows(min_row=2, max_row=53, min_col=1, max_col=5, values_only=True), start=1):
    for col_idx, value in enumerate(row):
        label = tk.Label(content_frame, text=value if value else "", borderwidth=1, relief="solid", padx=5, pady=5)
        label.grid(row=row_idx, column=col_idx, sticky="nsew", padx=2, pady=2)
    
    entry = tk.Entry(content_frame, width=15)
    entry.grid(row=row_idx, column=len(headers) - 1, padx=2, pady=2, sticky="nsew")

    def number_search(entry=entry, course_name=row[0], course_code=row[1], course_time=row[2]):
        student_id = entry.get()
        if not student_id:
            messagebox.showwarning("錯誤", "請輸入學號")
            return

        entry.delete(0, END)
        path = search(student_id)

        if path:
            total_credits = calculate_total_credits(path)
            course_credit = get_course_credit(course_name)

            if is_course_in_schedule(path, course_name):
                messagebox.showinfo("提醒", f"課程 {course_name} 已存在於課表中")
                return

            if total_credits + course_credit > 25:
                messagebox.showerror("超過學分上限", "加選失敗，超過學分上限！")
                return

            day, time_range = course_time.split()
            col_num, start_row, end_row = map_course_time_to_schedule(day, time_range)

            if start_row and end_row:
                # 檢查課程剩餘名額
                remaining_spots = get_course_remaining_spots(course_name)
                print(f"{remaining_spots}")
                if remaining_spots is not None:
                    if remaining_spots > 0:
                        add_course_to_schedule(path, course_name, course_code, col_num, start_row, end_row)
                        update_course_remaining_spots(course_name, remaining_spots - 1)  # 減少名額
                        display_schedule(path)
                    else:
                        messagebox.showerror("加選失敗", "修課人數已滿，加選失敗")
                else:
                    messagebox.showerror("錯誤", "無法取得課程剩餘名額")
            else:
                messagebox.showerror("錯誤", "無法對應課程時間")

        else:
            messagebox.showinfo("錯誤", "學號輸入錯誤")

    confirm_button = tk.Button(content_frame, text="確認", command=number_search)
    confirm_button.grid(row=row_idx, column=len(headers), padx=2, pady=2, sticky="nsew")

# 設定滾動區域範圍
content_frame.update_idletasks()
canvas.config(scrollregion=canvas.bbox("all"))

Swindow.mainloop()
